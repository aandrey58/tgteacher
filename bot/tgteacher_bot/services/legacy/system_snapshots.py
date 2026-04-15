import os
import time
import json
import datetime
from typing import Any, Dict, List, Optional, Tuple

try:
    import psutil  # type: ignore
except ImportError:
    psutil = None

from telegram.ext import ContextTypes
from zoneinfo import ZoneInfo
from tgteacher_bot.db.pool import get_pool
from tgteacher_bot.handlers.admin.admin_status import (
    get_db_ping,
    get_telegram_ping,
    get_pg_connections_count,
    get_avg_response_time,
    get_errors_last_24h,
)
from tgteacher_bot.db.families_repo import DB_ERRORS
from tgteacher_bot.handlers.admin.admin_status import ERRORS as BOT_ERRORS
from tgteacher_bot.core import paths

EXPORTS_DIR = str(paths.bot_dir() / 'exports')
AUTO_SNAPSHOT_INTERVAL_SECONDS = 5 * 60  # каждые 5 минут


async def init_system_snapshots_pg() -> None:
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute(
            '''
            CREATE TABLE IF NOT EXISTS system_snapshots (
                id BIGSERIAL PRIMARY KEY,
                ts TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                data JSONB NOT NULL
            );
            '''
        )
        await conn.execute(
            '''
            CREATE INDEX IF NOT EXISTS idx_system_snapshots_ts ON system_snapshots (ts DESC);
            '''
        )


async def _collect_metrics_core(context: ContextTypes.DEFAULT_TYPE) -> Dict[str, Any]:
    bot_start_time = context.application.bot_data.get('bot_start_time', time.time())

    metrics: Dict[str, Any] = {
        'ts': datetime.datetime.utcnow().isoformat() + 'Z',
        'uptime_sec': int(time.time() - bot_start_time),
        'ram_process_mb': None,
        'ram_used_mb': None,
        'ram_total_mb': None,
        'ram_free_mb': None,
        'swap_used_mb': None,
        'cpu_percent_avg': None,
        'cpu_percent_max_core': None,
        'disk_free_mb': None,
        'disk_total_mb': None,
        'pg_connections_count': None,
        'db_ping_ms': None,
        'tg_ping_ms': None,
        'avg_response_time_sec': None,
        'errors_last_1h': None,
        'errors_last_24h': None,
        'db_errors_last_24h': None,
        'sources': {
            'psutil': bool(psutil),
        }
    }

    # Статические поля (Python/OS) из снимков исключены — они есть в экране статуса

    # psutil-метрики
    if psutil:
        try:
            process = psutil.Process(os.getpid())
            mem_info = process.memory_info()
            metrics['ram_process_mb'] = mem_info.rss // (1024 * 1024)

            vm = psutil.virtual_memory()
            metrics['ram_total_mb'] = vm.total // (1024 * 1024)
            metrics['ram_used_mb'] = vm.used // (1024 * 1024)
            metrics['ram_free_mb'] = vm.available // (1024 * 1024)

            swap = psutil.swap_memory()
            metrics['swap_used_mb'] = swap.used // (1024 * 1024)

            # CPU: среднее и максимум по ядрам (без списка, экономим размер)
            cpu_avg = process.cpu_percent(interval=0.5)
            per_core = psutil.cpu_percent(percpu=True)
            metrics['cpu_percent_avg'] = cpu_avg
            metrics['cpu_percent_max_core'] = max(per_core) if per_core else cpu_avg

            disk = psutil.disk_usage('/')
            metrics['disk_free_mb'] = disk.free // (1024 * 1024)
            metrics['disk_total_mb'] = disk.total // (1024 * 1024)
        except Exception as e:
            metrics['sources']['psutil_error'] = str(e)

    # PG connections
    try:
        metrics['pg_connections_count'] = await get_pg_connections_count()
    except Exception as e:
        metrics['pg_connections_count'] = None
        metrics['sources']['pg_connections_error'] = str(e)

    # Pings
    try:
        metrics['db_ping_ms'] = await get_db_ping()
    except Exception as e:
        metrics['db_ping_ms'] = None
        metrics['sources']['db_ping_error'] = str(e)

    try:
        metrics['tg_ping_ms'] = await get_telegram_ping(context)
    except Exception as e:
        metrics['tg_ping_ms'] = None
        metrics['sources']['tg_ping_error'] = str(e)

    # Bot metrics
    try:
        metrics['avg_response_time_sec'] = get_avg_response_time()
    except Exception as e:
        metrics['avg_response_time_sec'] = None
        metrics['sources']['avg_response_time_error'] = str(e)

    try:
        now = time.time()
        errors_1h = len([t for t, _ in BOT_ERRORS if now - t < 3600])
        metrics['errors_last_1h'] = errors_1h
        metrics['errors_last_24h'] = get_errors_last_24h()
        # DB errors (здесь DB_ERRORS уже из data_db_pg)
        db_errors_24h = len([t for t, _ in DB_ERRORS if now - t < 86400])
        metrics['db_errors_last_24h'] = db_errors_24h
    except Exception as e:
        metrics['sources']['errors_calc_error'] = str(e)

    return metrics


async def capture_and_store_snapshot(context: ContextTypes.DEFAULT_TYPE) -> Dict[str, Any]:
    data = await _collect_metrics_core(context)
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute('INSERT INTO system_snapshots (data) VALUES ($1)', json.dumps(data))
    
    # Оптимизация: сборка мусора после создания снимка
    import gc
    gc.collect()
    
    return data


async def get_last_snapshot() -> Optional[Dict[str, Any]]:
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow('SELECT ts, data FROM system_snapshots ORDER BY ts DESC LIMIT 1')
        if not row:
            return None
        data = row['data'] if isinstance(row['data'], dict) else json.loads(row['data'])
        data['ts'] = row['ts'].isoformat()
        return data


async def get_prev_snapshot() -> Optional[Dict[str, Any]]:
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch('SELECT ts, data FROM system_snapshots ORDER BY ts DESC LIMIT 2')
        if len(rows) < 2:
            return None
        prev = rows[1]
        data = prev['data'] if isinstance(prev['data'], dict) else json.loads(prev['data'])
        data['ts'] = prev['ts'].isoformat()
        return data


async def get_snapshots_between(start: datetime.datetime, end: datetime.datetime) -> List[Dict[str, Any]]:
    # Нормализуем границы к UTC (tz-aware), чтобы избежать сюрпризов TZ
    if start.tzinfo is None:
        start = start.replace(tzinfo=datetime.timezone.utc)
    else:
        start = start.astimezone(datetime.timezone.utc)
    if end.tzinfo is None:
        end = end.replace(tzinfo=datetime.timezone.utc)
    else:
        end = end.astimezone(datetime.timezone.utc)
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch('SELECT ts, data FROM system_snapshots WHERE ts >= $1 AND ts <= $2 ORDER BY ts DESC', start, end)
        result: List[Dict[str, Any]] = []
        for r in rows:
            d = r['data'] if isinstance(r['data'], dict) else json.loads(r['data'])
            d['ts'] = r['ts'].isoformat()
            result.append(d)
        return result


async def get_recent_snapshots(limit: int = 20) -> List[Dict[str, Any]]:
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch('SELECT ts, data FROM system_snapshots ORDER BY ts DESC LIMIT $1', limit)
        result: List[Dict[str, Any]] = []
        for r in rows:
            d = r['data'] if isinstance(r['data'], dict) else json.loads(r['data'])
            d['ts'] = r['ts'].isoformat()
            result.append(d)
        return result


def _format_snapshot_text(data: Dict[str, Any]) -> str:
    # Компактный рендер без лишних подробностей
    uptime = data.get('uptime_sec', 0)
    uptime_str = f"{uptime//3600}ч {(uptime%3600)//60}м {uptime%60}с"
    # Локализованное время снимка (МСК), если доступно
    ts_line = None
    ts_raw = data.get('ts')
    if isinstance(ts_raw, str):
        try:
            dt = datetime.datetime.fromisoformat(ts_raw.replace('Z', '+00:00'))
            dt_msk = dt.astimezone(ZoneInfo('Europe/Moscow'))
            ts_line = f"🕒 <b>Время снимка (МСК):</b> {dt_msk.strftime('%d-%m-%Y %H:%M:%S')}"
        except Exception:
            ts_line = None
    lines = [
        '📸 <b>Снимок системы</b>\n',
        *( [ts_line] if ts_line else [] ),
        f"⏱ <b>Uptime:</b> {uptime_str}",
        f"🧠 <b>CPU ср/макс ядро:</b> {data.get('cpu_percent_avg', '-')}% / {data.get('cpu_percent_max_core', '-') }%",
        f"💾 <b>RAM процесса:</b> {data.get('ram_process_mb', '-')} MB",
        f"💾 <b>RAM всего:</b> {data.get('ram_used_mb', '-')}/{data.get('ram_total_mb', '-')} MB (свободно: {data.get('ram_free_mb', '-')} MB)",
        f"💽 <b>Диск:</b> {data.get('disk_free_mb', '-')}/{data.get('disk_total_mb', '-')} MB свободно",
        f"🗄 <b>PG соединений:</b> {data.get('pg_connections_count', '-')}",
        f"🏓 <b>Пинг БД/ТГ:</b> {data.get('db_ping_ms','-')} / {data.get('tg_ping_ms','-')} мс",
        f"⏳ <b>Среднее время ответа:</b> {data.get('avg_response_time_sec','-')} сек",
        f"❗ <b>Ошибок 1ч/24ч (БД 24ч):</b> {data.get('errors_last_1h','-')} / {data.get('errors_last_24h','-')} ({data.get('db_errors_last_24h','-')})",
    ]
    return '\n'.join(lines)


def _arrow(delta: float) -> str:
    if abs(delta) < 1e-9:
        return '→'
    return '↑' if delta > 0 else '↓'


def _fmt_trend_line(label: str, curr: Any, prev: Any, unit: str, precision: int, threshold: float, emoji: str = '') -> Optional[str]:
    if not isinstance(curr, (int, float)) or not isinstance(prev, (int, float)):
        return None
    delta = float(curr) - float(prev)
    if abs(delta) < threshold:
        return None
    arrow = '↑' if delta > 0 else '↓'
    val = round(abs(delta), precision)
    sign = '+' if delta > 0 else '−'
    # Неразрывный пробел перед единицей измерения для красоты
    space_unit = unit if unit == '' else f"{unit}"
    return f"- {emoji} {label}: {arrow} {sign}{val}{space_unit}".strip()


def format_snapshot_text_with_trends(curr: Dict[str, Any], prev: Optional[Dict[str, Any]]) -> str:
    base = _format_snapshot_text(curr)
    if not prev:
        return base
    lines: List[str] = ["", "📈 <b>Тренд к предыдущему</b>"]
    # Пороговые значения для отсечения шума
    cpu_line = _fmt_trend_line('CPU', curr.get('cpu_percent_avg'), prev.get('cpu_percent_avg'), '%', 1, threshold=0.1, emoji='🧠')
    ram_line = _fmt_trend_line('RAM процесса', curr.get('ram_process_mb'), prev.get('ram_process_mb'), ' MB', 0, threshold=1.0, emoji='💾')
    db_line = _fmt_trend_line('Пинг БД', curr.get('db_ping_ms'), prev.get('db_ping_ms'), ' мс', 0, threshold=1.0, emoji='🏓')
    tg_line = _fmt_trend_line('Пинг TG', curr.get('tg_ping_ms'), prev.get('tg_ping_ms'), ' мс', 0, threshold=10.0, emoji='🏓')
    rt_line = _fmt_trend_line('Время ответа', curr.get('avg_response_time_sec'), prev.get('avg_response_time_sec'), ' с', 2, threshold=0.05, emoji='⏳')
    pg_line = _fmt_trend_line('PG соединения', curr.get('pg_connections_count'), prev.get('pg_connections_count'), '', 0, threshold=1.0, emoji='🗄')

    for ln in [cpu_line, ram_line, db_line, tg_line, rt_line, pg_line]:
        if ln:
            lines.append(ln)

    if len(lines) == 2:
        lines.append('- без изменений')
    return base + '\n' + '\n'.join(lines)


async def export_snapshots_csv(period: str) -> Optional[str]:
    now = datetime.datetime.now(datetime.timezone.utc)
    if period == '1h':
        start = now - datetime.timedelta(hours=1)
    elif period == '24h':
        start = now - datetime.timedelta(hours=24)
    elif period == '7d':
        start = now - datetime.timedelta(days=7)
    else:
        start = now - datetime.timedelta(hours=24)

    snapshots = await get_snapshots_between(start, now)
    if not snapshots:
        return None

    if not os.path.exists(EXPORTS_DIR):
        os.makedirs(EXPORTS_DIR)

    safe_period = period.replace('/', '_')
    file_path = os.path.join(EXPORTS_DIR, f'system_snapshots_{safe_period}_{now.strftime("%Y-%m-%d_%H-%M-%S")}.csv')

    # Заголовки (RU) с маппингом ключ -> название
    fields_map = [
        ('ts', 'Время (МСК)'),
        ('uptime_sec', 'Аптайм (сек)'),
        ('ram_process_mb', 'RAM процесса (MB)'),
        ('ram_used_mb', 'RAM использовано (MB)'),
        ('ram_total_mb', 'RAM всего (MB)'),
        ('ram_free_mb', 'RAM свободно (MB)'),
        ('swap_used_mb', 'Swap (MB)'),
        ('cpu_percent_avg', 'CPU среднее (%)'),
        ('cpu_percent_max_core', 'CPU макс ядро (%)'),
        ('disk_free_mb', 'Диск свободно (MB)'),
        ('disk_total_mb', 'Диск всего (MB)'),
        ('pg_connections_count', 'PG соединений'),
        ('db_ping_ms', 'Пинг БД (мс)'),
        ('tg_ping_ms', 'Пинг Telegram (мс)'),
        ('avg_response_time_sec', 'Время ответа (сек)'),
        ('errors_last_1h', 'Ошибок за 1ч'),
        ('errors_last_24h', 'Ошибок за 24ч'),
        ('db_errors_last_24h', 'Ошибок БД за 24ч'),
    ]

    import csv
    with open(file_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow([header for _, header in fields_map])
        msk_tz = ZoneInfo('Europe/Moscow')
        for s in snapshots:
            # Преобразуем ts в МСК и формат ДД-ММ-ГГГГ ЧЧ:ММ:СС
            ts_raw = s.get('ts')
            ts_val = ''
            if isinstance(ts_raw, str):
                try:
                    dt = datetime.datetime.fromisoformat(ts_raw.replace('Z', '+00:00'))
                    ts_val = dt.astimezone(msk_tz).strftime('%d-%m-%Y %H:%M:%S')
                except Exception:
                    ts_val = ts_raw
            values_map = {
                'ts': ts_val,
                'uptime_sec': s.get('uptime_sec'),
                'ram_process_mb': s.get('ram_process_mb'),
                'ram_used_mb': s.get('ram_used_mb'),
                'ram_total_mb': s.get('ram_total_mb'),
                'ram_free_mb': s.get('ram_free_mb'),
                'swap_used_mb': s.get('swap_used_mb'),
                'cpu_percent_avg': s.get('cpu_percent_avg'),
                'cpu_percent_max_core': s.get('cpu_percent_max_core'),
                'disk_free_mb': s.get('disk_free_mb'),
                'disk_total_mb': s.get('disk_total_mb'),
                'pg_connections_count': s.get('pg_connections_count'),
                'db_ping_ms': s.get('db_ping_ms'),
                'tg_ping_ms': s.get('tg_ping_ms'),
                'avg_response_time_sec': s.get('avg_response_time_sec'),
                'errors_last_1h': s.get('errors_last_1h'),
                'errors_last_24h': s.get('errors_last_24h'),
                'db_errors_last_24h': s.get('db_errors_last_24h'),
            }
            writer.writerow([values_map.get(key, '') for key, _ in fields_map])

    return file_path


async def export_snapshots_xlsx(period: str) -> Optional[str]:
    now = datetime.datetime.now(datetime.timezone.utc)
    if period == '1h':
        start = now - datetime.timedelta(hours=1)
    elif period == '24h':
        start = now - datetime.timedelta(hours=24)
    elif period == '7d':
        start = now - datetime.timedelta(days=7)
    else:
        start = now - datetime.timedelta(hours=24)

    snapshots = await get_snapshots_between(start, now)
    if not snapshots:
        return None

    if not os.path.exists(EXPORTS_DIR):
        os.makedirs(EXPORTS_DIR)

    safe_period = period.replace('/', '_')
    file_path = os.path.join(EXPORTS_DIR, f'system_snapshots_{safe_period}_{now.strftime("%Y-%m-%d_%H-%M-%S")}.xlsx')

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = 'Снимки системы'

    headers = [
        'Время (МСК)', 'Аптайм (сек)', 'RAM процесса (MB)', 'RAM использовано (MB)', 'RAM всего (MB)',
        'RAM свободно (MB)', 'Swap (MB)', 'CPU среднее (%)', 'CPU макс ядро (%)',
        'Диск свободно (MB)', 'Диск всего (MB)', 'PG соединений',
        'Пинг БД (мс)', 'Пинг Telegram (мс)', 'Время ответа (сек)',
        'Ошибок за 1ч', 'Ошибок за 24ч', 'Ошибок БД за 24ч'
    ]

    ws.append(headers)

    msk_tz = ZoneInfo('Europe/Moscow')
    center = Alignment(horizontal='center', vertical='center')
    thin = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
                  top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'))

    for s in snapshots:
        ts_raw = s.get('ts')
        ts_val = ''
        if isinstance(ts_raw, str):
            try:
                dt = datetime.datetime.fromisoformat(ts_raw.replace('Z', '+00:00'))
                ts_val = dt.astimezone(msk_tz).strftime('%d-%m-%Y %H:%M:%S')
            except Exception:
                ts_val = ts_raw
        row = [
            ts_val,
            s.get('uptime_sec'),
            s.get('ram_process_mb'),
            s.get('ram_used_mb'),
            s.get('ram_total_mb'),
            s.get('ram_free_mb'),
            s.get('swap_used_mb'),
            s.get('cpu_percent_avg'),
            s.get('cpu_percent_max_core'),
            s.get('disk_free_mb'),
            s.get('disk_total_mb'),
            s.get('pg_connections_count'),
            s.get('db_ping_ms'),
            s.get('tg_ping_ms'),
            s.get('avg_response_time_sec'),
            s.get('errors_last_1h'),
            s.get('errors_last_24h'),
            s.get('db_errors_last_24h'),
        ]
        ws.append(row)

    # Стилизация: центр и границы, заголовки жирные
    for col in ws.columns:
        for cell in col:
            cell.alignment = center
            cell.border = thin
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Примерные ширины
    widths = [19, 14, 18, 20, 16, 18, 12, 17, 18, 18, 16, 15, 14, 18, 20, 14, 16, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    wb.save(file_path)
    wb.close()
    return file_path


# Джоба автоснимков и ретеншна
async def auto_snapshot_job(context: ContextTypes.DEFAULT_TYPE):
    try:
        if not context.application.bot_data.get('snapshots_auto_enabled', True):
            return
        await capture_and_store_snapshot(context)
    except Exception as e:
        # Игнорируем, чтобы не валить джобу
        pass


async def retention_job(context: ContextTypes.DEFAULT_TYPE):
    # Удаляем снимки старше 30 дней (лёгкая ретеншн-политика)
    try:
        cutoff = datetime.datetime.utcnow() - datetime.timedelta(days=30)
        pool = await get_pool()
        async with pool.acquire() as conn:
            await conn.execute('DELETE FROM system_snapshots WHERE ts < $1', cutoff)
    except Exception:
        pass


def format_snapshot_text_for_msg(data: Dict[str, Any]) -> str:
    return _format_snapshot_text(data) 