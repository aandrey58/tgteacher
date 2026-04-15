import pandas as pd
import os
import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.formatting.rule import CellIsRule
from tgteacher_bot.db.user_repo import get_pool
from tgteacher_bot.db.families_repo import get_stage1_words_pg, get_stage2_tasks_pg, get_stage3_tasks_pg, get_stage6_tasks_pg, get_stage8_tasks_pg, get_stage7_tasks_pg
from tgteacher_bot.db.user_repo import get_all_users_progress_for_family_pg, get_all_stage_answers_for_family_pg, get_completed_tasks_pg, get_family_total_completion_count_pg
from tgteacher_bot.db.families_repo import get_family_data_pg
import re
from tgteacher_bot.db.families_repo import get_stage4_tasks_pg
from tgteacher_bot.db.families_repo import get_stage5_tasks_pg
from datetime import datetime, timedelta
import asyncio
from tgteacher_bot.core import paths

EXPORTS_DIR = str(paths.bot_dir() / 'exports')

def get_period_dates(period):
    """Получение дат для периода"""
    now = datetime.now()
    if period == 'today':
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end = now
    elif period == 'week':
        start = now - timedelta(days=7)
        end = now
    elif period == 'month':
        start = now - timedelta(days=30)
        end = now
    elif period.startswith('custom_'):
        try:
            _, start_str, end_str = period.split('_')
            start = datetime.strptime(start_str, '%Y-%m-%d')
            end = datetime.strptime(end_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59, microsecond=999999)
            return start, end
        except (ValueError, IndexError):
            return None, None
    else:
        start = None
        end = None
    return start, end

def _prepare_users_df(users):
    """Готовит DataFrame пользователей для экспорта."""
    if not users:
        return pd.DataFrame()
    df = pd.DataFrame(users)
    if 'registered_at' in df:
        df['registered_at'] = pd.to_datetime(df['registered_at']).dt.strftime('%Y-%m-%d %H:%M:%S')
    if 'is_subscribed' in df:
        df['is_subscribed'] = df['is_subscribed'].map({True: 'ДА', False: 'НЕТ'})
    
    columns_map = {
        'user_id': 'Telegram ID',
        'username': 'Username',
        'first_name': 'Имя',
        'last_name': 'Фамилия',
        'registered_at': 'Дата регистрации',
        'is_subscribed': 'Подписка',
        'subscription_count': 'Кол-во подписок',
        'last_active_at': 'Последнее взаимодействие',
    }
    df = df.rename(columns=columns_map)
    # Убедимся, что все колонки на месте, даже если данных не было
    final_columns = [
        'Telegram ID', 'Username', 'Имя', 'Фамилия', 
        'Дата регистрации', 'Подписка', 'Кол-во подписок', 'Последнее взаимодействие'
    ]
    for col in final_columns:
        if col not in df.columns:
            df[col] = None
    # Форматируем дату последнего взаимодействия
    if 'Последнее взаимодействие' in df:
        df['Последнее взаимодействие'] = pd.to_datetime(df['Последнее взаимодействие']).dt.strftime('%Y-%m-%d %H:%M:%S')
    return df[final_columns]

def _style_excel_sheet(ws):
    """Применяет стили к листу Excel."""
    col_width = 18.7
    thin_border = Border(left=Side(style="thin", color="000000"),
                         right=Side(style="thin", color="000000"),
                         top=Side(style="thin", color="000000"),
                         bottom=Side(style="thin", color="000000"))
    center_align = Alignment(horizontal='center', vertical='center')

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = col_width
        for cell in col:
            cell.alignment = center_align
            cell.border = thin_border

    # MCP: Делаем заголовки жирными
    for cell in ws[1]:
        cell.font = Font(bold=True)

async def _export_df_to_excel(df, file_path):
    """Сохраняет DataFrame в Excel и применяет стили."""
    if not os.path.exists(EXPORTS_DIR):
        os.makedirs(EXPORTS_DIR)
    
    df.to_excel(file_path, index=False)
    
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    _style_excel_sheet(ws)
    # MCP: Устанавливаем ширину для столбца "Последнее взаимодействие" (H)
    ws.column_dimensions['H'].width = 27
    wb.save(file_path)
    wb.close()
    return file_path

async def export_all_users_to_excel():
    pool = await get_pool()
    now_str = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    file_path = os.path.join(EXPORTS_DIR, f'all_users_export_{now_str}.xlsx')

    if not os.path.exists(EXPORTS_DIR):
        os.makedirs(EXPORTS_DIR)

    # Создаем новую книгу и лист
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Все пользователи"

    # Записываем заголовки
    headers = [
        'Telegram ID', 'Username', 'Имя', 'Фамилия',
        'Дата регистрации', 'Подписка', 'Кол-во подписок', 'Последнее взаимодействие'
    ]
    ws.append(headers)

    # Записываем данные построчно
    async with pool.acquire() as conn:
        async with conn.transaction(): # MCP: Добавил транзакцию
            async for record in conn.cursor('SELECT user_id, username, first_name, last_name, registered_at, is_subscribed, subscription_count, last_active_at FROM users ORDER BY registered_at DESC'):
                row_data = [
                    record['user_id'],
                    record['username'] or '-',
                    record['first_name'] or '-',
                    record['last_name'] or '-',
                    record['registered_at'].strftime('%Y-%m-%d %H:%M:%S') if record['registered_at'] else '-',
                    'ДА' if record['is_subscribed'] else 'НЕТ',
                    record['subscription_count'],
                    record['last_active_at'].strftime('%Y-%m-%d %H:%M:%S') if record['last_active_at'] else '-',
                ]
                ws.append(row_data)

    # Применяем стили к листу
    _style_excel_sheet(ws)
    ws.column_dimensions['H'].width = 27 # Ширина для колонки "Последнее взаимодействие"

    wb.save(file_path)
    wb.close()
    return file_path

async def export_new_users_to_excel(period):
    pool = await get_pool()
    start, end = get_period_dates(period)
    
    now_str = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    file_path = os.path.join(EXPORTS_DIR, f'{period}_new_users_{now_str}.xlsx')

    if not os.path.exists(EXPORTS_DIR):
        os.makedirs(EXPORTS_DIR)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Новые пользователи"

    headers = [
        'Telegram ID', 'Username', 'Имя', 'Фамилия',
        'Дата регистрации', 'Подписка', 'Кол-во подписок', 'Последнее взаимодействие'
    ]
    ws.append(headers)

    async with pool.acquire() as conn:
        async with conn.transaction(): # MCP: Добавил транзакцию
            async for record in conn.cursor('SELECT user_id, username, first_name, last_name, registered_at, is_subscribed, subscription_count, last_active_at FROM users WHERE registered_at >= $1 AND registered_at <= $2 ORDER BY registered_at DESC', start, end):
                row_data = [
                    record['user_id'],
                    record['username'] or '-',
                    record['first_name'] or '-',
                    record['last_name'] or '-',
                    record['registered_at'].strftime('%Y-%m-%d %H:%M:%S') if record['registered_at'] else '-',
                    'ДА' if record['is_subscribed'] else 'НЕТ',
                    record['subscription_count'],
                    record['last_active_at'].strftime('%Y-%m-%d %H:%M:%S') if record['last_active_at'] else '-',
                ]
                ws.append(row_data)

    _style_excel_sheet(ws)
    ws.column_dimensions['H'].width = 27

    wb.save(file_path)
    wb.close()
    return file_path

async def _get_family_progress_stats_for_excel(family_id):
    from tgteacher_bot.db.user_repo import get_all_users_progress_for_family_pg, get_family_stage_answers_stats_pg, get_family_finished_users_pg, get_family_total_completion_count_pg
    from tgteacher_bot.db.families_repo import get_all_stage_tasks_counts_for_families
    from collections import defaultdict
    # Получаем данные
    all_user_progress = await get_all_users_progress_for_family_pg(family_id)
    answers_stats = await get_family_stage_answers_stats_pg(family_id)
    total_tasks_per_stage = (await get_all_stage_tasks_counts_for_families([family_id])).get(family_id, {})
    finished_user_ids = set(await get_family_finished_users_pg(family_id))
    # MCP: Получаем общее количество завершений семьи
    total_completion_count = await get_family_total_completion_count_pg(family_id)
    total_stages_count = sum(1 for count in total_tasks_per_stage.values() if count > 0)

    users_data = defaultdict(lambda: {'completed_stages': set(), 'current_stage': 0, 'total_tasks_done': 0})
    for progress in all_user_progress:
        user_id = progress['user_id']
        stage_num = progress['stage_num']
        completed_tasks = progress['completed_tasks']
        users_data[user_id]['total_tasks_done'] += completed_tasks
        if completed_tasks >= total_tasks_per_stage.get(stage_num, 0):
            users_data[user_id]['completed_stages'].add(stage_num)
        if stage_num > users_data[user_id]['current_stage']:
            users_data[user_id]['current_stage'] = stage_num

    completed_users_count = 0
    in_progress_users_count = 0
    finished_users_count = 0
    users_on_stage = defaultdict(int)
    for user_id, data in users_data.items():
        is_fully_completed = len(data['completed_stages']) == total_stages_count
        is_finished = user_id in finished_user_ids
        if is_fully_completed:
            completed_users_count += 1
        if is_finished:
            finished_users_count += 1
        if not is_fully_completed and not is_finished:
            in_progress_users_count += 1
            current_stage = data['current_stage']
            if current_stage > 0:
                users_on_stage[current_stage] += 1

    stage_errors = defaultdict(lambda: {'correct': 0, 'incorrect': 0})
    for stat in answers_stats:
        stage_num = stat['stage_num']
        count = stat['count']
        if stat['is_correct']:
            stage_errors[stage_num]['correct'] += count
        else:
            stage_errors[stage_num]['incorrect'] += count

    analytics = []
    analytics.append(["Аналитика по группе слов:", None])
    analytics.append(["Всего пользователей", len(users_data)])
    analytics.append(["Прошли полностью", completed_users_count])
    analytics.append(["Завершили", finished_users_count])
    analytics.append(["В процессе", in_progress_users_count])
    # MCP: Добавляем общее количество завершений
    analytics.append(["Всего прохождений", total_completion_count])
    analytics.append(["", ""])  # пустая строка
    analytics.append(["Распределение по этапам", None])
    # MCP: добавил этап 5 в распределение
    for stage in [1, 2, 3, 4, 5, 6, 7, 8]:
        analytics.append([f"  Этап {stage}", users_on_stage.get(stage, 0)])
    analytics.append(["", ""])  # пустая строка
    analytics.append(["Ошибки по этапам", None])
    total_errors = 0
    total_answers = 0
    hardest_stage = None
    hardest_percent = 0
    easiest_stage = None
    easiest_percent = 100
    # MCP: добавил этап 5 в расчёт ошибок
    for stage_num in [2, 3, 4, 5, 6, 7, 8]:
        stats = stage_errors[stage_num]
        stage_total = stats['correct'] + stats['incorrect']
        err_percent = (stats['incorrect'] / stage_total * 100) if stage_total > 0 else 0
        analytics.append([f"  Этап {stage_num}", f"{err_percent:.1f}% ошибок ({stats['incorrect']} из {stage_total})"])
        total_errors += stats['incorrect']
        total_answers += stage_total
        if stage_total > 0 and err_percent > hardest_percent:
            hardest_stage = stage_num
            hardest_percent = err_percent
        if stage_total > 0 and err_percent < easiest_percent:
            easiest_stage = stage_num
            easiest_percent = err_percent
    analytics.append(["", ""])  # пустая строка
    avg_error_percent = (total_errors / total_answers * 100) if total_answers > 0 else 0
    analytics.append(["Всего ошибок", total_errors])
    analytics.append(["Всего ответов", total_answers])
    analytics.append(["Средний процент ошибок", f"{avg_error_percent:.1f}%"])
    if hardest_stage is not None:
        analytics.append(["Самый сложный этап", f"Этап {hardest_stage} ({hardest_percent:.1f}%)"])
    if easiest_stage is not None:
        analytics.append(["Самый лёгкий этап", f"Этап {easiest_stage} ({easiest_percent:.1f}%)"])
    return analytics

async def export_family_progress_to_excel(family_id: int):
    # 1. Получаем структуру группы слов
    s1_tasks = await get_stage1_words_pg(family_id)
    s2_tasks = await get_stage2_tasks_pg(family_id)
    s3_tasks = await get_stage3_tasks_pg(family_id)
    s4_tasks = await get_stage4_tasks_pg(family_id)
    # MCP: добавил получение задач этапа 5
    s5_tasks = await get_stage5_tasks_pg(family_id)
    s6_tasks = await get_stage6_tasks_pg(family_id)
    s7_tasks = await get_stage7_tasks_pg(family_id)
    s8_tasks = await get_stage8_tasks_pg(family_id)

    # MCP: Получаем имя группы слов для файла
    family_meta = await get_family_data_pg(family_id)
    family_name = family_meta['name'] if family_meta and 'name' in family_meta else f'family_{family_id}'
    # Очищаем имя для файла (только латиница, цифры, _)
    safe_family_name = re.sub(r'[^A-Za-z0-9_]', '_', family_name)

    # 2. Собираем список всех пользователей, кто делал хоть что-то в этой группе слов
    pool = await get_pool()
    async with pool.acquire() as conn:
        user_rows = await conn.fetch('SELECT DISTINCT user_id FROM user_task_progress WHERE family_idx = $1', family_id)
        user_ids = [row['user_id'] for row in user_rows]

    # 3. Формируем заголовки столбцов
    columns = ['ID']
    stage_task_map = []  # (stage_num, task_idx, col_name)
    for stage_num, tasks, label in [
        (1, s1_tasks, 'Этап 1'), (2, s2_tasks, 'Этап 2'), (3, s3_tasks, 'Этап 3'),
        (4, s4_tasks, 'Этап 4'), (5, s5_tasks, 'Этап 5'), (6, s6_tasks, 'Этап 6'), (7, s7_tasks, 'Этап 7'), (8, s8_tasks, 'Этап 8'),
    ]:
        for idx, _ in enumerate(tasks):
            col_name = f'{stage_num}.{idx+1}'
            columns.append(col_name)
            stage_task_map.append((stage_num, idx, col_name))

    # 4. Вытаскиваем все ответы одним махом
    all_completed_s1 = await get_completed_tasks_pg(user_ids, family_id, 1) # Предполагаем, что функция может принять список user_id
    all_answers = await get_all_stage_answers_for_family_pg(user_ids, family_id) # И эта тоже

    # 5. Собираем данные по каждому пользователю в памяти
    data = []
    for user_id in user_ids:
        row = [user_id]
        user_completed_s1 = all_completed_s1.get(user_id, set())
        user_answers = all_answers.get(user_id, {})

        for stage_num, idx, _ in stage_task_map:
            if stage_num == 1:
                row.append('Верно' if idx in user_completed_s1 else '—')
            else:
                if idx in user_answers.get(stage_num, {}):
                    _, is_correct = user_answers[stage_num][idx]
                    row.append('Верно' if is_correct else 'Неверно')
                else:
                    row.append('—')
        data.append(row)

    # 6. Создаём DataFrame и сохраняем в Excel
    if not os.path.exists(EXPORTS_DIR):
        os.makedirs(EXPORTS_DIR)
    now_str = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    file_path = os.path.join(EXPORTS_DIR, f'family_{family_id}_{safe_family_name}_progress_{now_str}.xlsx')
    df = pd.DataFrame(data, columns=columns)
    df.to_excel(file_path, index=False)

    # 7. Красим ячейки через условное форматирование
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    ws.column_dimensions['A'].width = 24
    
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    green_font = Font(color='006100')
    red_font = Font(color='9C0006')

    ws.conditional_formatting.add(f'B2:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}',
                                 CellIsRule(operator='equal', formula=['"Верно"'], fill=green_fill, font=green_font))
    ws.conditional_formatting.add(f'B2:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}',
                                 CellIsRule(operator='equal', formula=['"Неверно"'], fill=red_fill, font=red_font))

    thin_border = Border(left=Side(style="thin", color="000000"),
                         right=Side(style="thin", color="000000"),
                         top=Side(style="thin", color="000000"),
                         bottom=Side(style="thin", color="000000"))
    center_align = Alignment(horizontal='center', vertical='center')

    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_align

    # MCP: === АНАЛИТИКА В КОНЕЦ EXCEL (async, без run_until_complete) ===
    analytics = await _get_family_progress_stats_for_excel(family_id)
    start_row = ws.max_row + 2
    for i, (k, v) in enumerate(analytics):
        ws.cell(row=start_row + i, column=1, value=k)
        ws.cell(row=start_row + i, column=2, value=v)
        ws.cell(row=start_row + i, column=1).font = Font(bold=True) if v is None else Font(bold=False)
        ws.cell(row=start_row + i, column=1).alignment = Alignment(horizontal='left')
        ws.cell(row=start_row + i, column=2).alignment = Alignment(horizontal='left')

    wb.save(file_path)
    wb.close()
    return file_path 

async def export_payment_history_to_excel(period):
    """Экспорт истории покупок за период"""
    
    pool = await get_pool()
    start, end = get_period_dates(period)
    
    now_str = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    file_path = os.path.join(EXPORTS_DIR, f'{period}_payment_history_{now_str}.xlsx')

    if not os.path.exists(EXPORTS_DIR):
        os.makedirs(EXPORTS_DIR)

    # Перед выгрузкой: обновляем статусы платежей из ЮKassa по платежам за период
    try:
        from tgteacher_bot.services.payments.yookassa_payment import get_yookassa_payment
        payment_api = get_yookassa_payment()
        async with pool.acquire() as conn:
            async with conn.transaction():
                # Берем только платежи со статусами, которые могут измениться
                refresh_query = '''
                    SELECT payment_id
                    FROM payments
                    WHERE created_at >= $1 AND created_at <= $2
                      AND status IN ('pending', 'waiting_for_capture', 'canceled')
                '''
                payment_ids = [r['payment_id'] for r in await conn.fetch(refresh_query, start, end)]
        # Параллельно дергаем YooKassa для обновления статусов, но без падений при ошибках
        await asyncio.gather(*[
            payment_api.check_payment_status(pid) for pid in payment_ids
        ], return_exceptions=True)
    except Exception:
        # Тихо игнорируем, чтобы выгрузка всё равно произошла
        pass

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "История покупок"

    headers = [
        'OrderID', 'Telegram ID', 'Username', 'Имя', 'Фамилия',
        'Дата оплаты', 'Срок подписки (мес)', 'Сумма (₽)', 'Статус платежа',
        'Подписка активна до'
    ]
    ws.append(headers)

    # МСК (UTC+3)
    from datetime import timezone
    moscow_tz = timezone(timedelta(hours=3))

    def to_moscow_str(dt, fmt):
        if not dt:
            return '-'
        try:
            if getattr(dt, 'tzinfo', None):
                dt_msk = dt.astimezone(moscow_tz)
            else:
                # Если таймзоны нет, считаем что это уже МСК (как и в остальных местах бота)
                dt_msk = dt.replace(tzinfo=moscow_tz)
            return dt_msk.strftime(fmt)
        except Exception:
            return '-'

    async with pool.acquire() as conn:
        async with conn.transaction():
            # Получаем все платежи за период с данными пользователей и расчётом даты окончания по каждому платежу
            query = '''
                SELECT 
                    p.payment_id,
                    p.user_id,
                    p.months,
                    p.amount,
                    p.status,
                    p.created_at as payment_date,
                    COALESCE(p.result_subscription_until, (p.created_at + make_interval(months => p.months))) AS active_until,
                    u.username,
                    u.first_name,
                    u.last_name
                FROM payments p
                LEFT JOIN users u ON p.user_id = u.user_id
                WHERE p.created_at >= $1 AND p.created_at <= $2
                ORDER BY p.created_at DESC
            '''
            async for record in conn.cursor(query, start, end):
                # Конвертируем сумму из копеек в рубли
                amount_rub = record['amount'] / 100 if record['amount'] else 0
                
                # Форматируем даты в МСК
                payment_date_str = to_moscow_str(record['payment_date'], '%Y-%m-%d %H:%M:%S')
                active_until_str = to_moscow_str(record['active_until'], '%d.%m.%Y %H:%M')
                
                row_data = [
                    record['payment_id'],
                    record['user_id'],
                    record['username'] or '-',
                    record['first_name'] or '-',
                    record['last_name'] or '-',
                    payment_date_str,
                    record['months'],
                    f"{amount_rub:.2f}",
                    record['status'],
                    active_until_str,
                ]
                ws.append(row_data)

    _style_excel_sheet(ws)
    
    # Устанавливаем ширину для столбцов с датами и OrderID
    ws.column_dimensions['A'].width = 36  # OrderID
    ws.column_dimensions['F'].width = 20  # Дата оплаты
    ws.column_dimensions['J'].width = 20  # Подписка активна до

    wb.save(file_path)
    wb.close()
    return file_path 